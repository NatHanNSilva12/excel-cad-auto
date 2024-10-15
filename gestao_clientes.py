import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
import pathlib
import matplotlib.pyplot as plt
import os

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearence()
        self.todo_sistema()

    def layout_config(self):
        self.title("Sistema de Gestão de Clientes")
        self.geometry("800x600")  # Aumentar o tamanho da tela

    def appearence(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=['#000', '#fff']).place(x=50, y=530)
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Light", "Dark", "System"], command=self.change_apm).place(x=50, y=560)

    def todo_sistema(self):
        frame = ctk.CTkFrame(self, width=800, height=50, corner_radius=0, bg_color="teal", fg_color="teal").place(x=0, y=10)
        title = ctk.CTkLabel(frame, text="Sistema de Gestão de Clientes", font=("Century Gothic bold", 24),
        text_color="#fff").place(x=190,y=10)

        span = ctk.CTkLabel(frame, text="Por favor, preencha todos os dados do cliente!", font=("Century Gothic bold", 16),
        text_color=["#000", "#fff"]).place(x=50, y=70)

        # Obtenha o caminho da área de trabalho do usuário
        desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')

        # Caminho completo para o arquivo Excel na área de trabalho
        excel_file_path = os.path.join(desktop_path, "Clientes.xlsx")

        ficheiro = pathlib.Path(excel_file_path)

        if ficheiro.exists():
            pass
        else:
            ficheiro = Workbook()
            folha = ficheiro.active
            folha['A1'] = "Nome completo"
            folha['B1'] = "Contato"
            folha['C1'] = "Código do Imóvel"
            folha['D1'] = "Endereço"
            folha['E1'] = "Observações"
            folha['F1'] = "Valor do Imóvel"
            folha['G1'] = "Status"
            ficheiro.save(excel_file_path)

        def submit():
            name = name_value.get()
            contact = contact_value.get()
            code = code_value.get()
            address = address_value.get()
            obs = obs_entry.get("1.0", END)
            value = value_entry.get()
            status = status_var.get()

            if name == "" or contact == "" or code == "" or address == "" or obs == "" or value == "":
                messagebox.showerror("Sistema", "ERRO!\nPor favor, preencha todos os dados!")
            else:
                ficheiro = openpyxl.load_workbook(excel_file_path)
                folha = ficheiro.active

                nova_linha = folha.max_row + 1
                folha.cell(column=1, row=nova_linha, value=name)
                folha.cell(column=2, row=nova_linha, value=contact)
                folha.cell(column=3, row=nova_linha, value=code)
                folha.cell(column=4, row=nova_linha, value=address)
                folha.cell(column=5, row=nova_linha, value=obs)
                folha.cell(column=6, row=nova_linha, value=value)
                
                # Salvar status com cor
                status_cell = folha.cell(column=7, row=nova_linha, value=status)
                if status == "Vendido":
                    status_cell.fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                elif status == "Encerrado":
                    status_cell.fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                elif status == "Em Processo":
                    status_cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                ficheiro.save(excel_file_path)
                messagebox.showinfo("Sistema", "Dados salvos com sucesso!")

        def clear():
            name_value.set("")
            contact_value.set("")
            code_value.set("")
            address_value.set("")
            obs_entry.delete(0.0, END)
            value_entry.delete(0, END)
            status_var.set("")

        def calcular_comissoes():
            # Carregar o Excel e calcular comissão de 6% sobre os imóveis vendidos
            ficheiro = openpyxl.load_workbook(excel_file_path)
            folha = ficheiro.active
            valores_vendidos = []
            nomes_clientes = []

            for row in folha.iter_rows(min_row=2, max_row=folha.max_row, values_only=True):
                if row[6] == "Vendido":
                    valores_vendidos.append(float(row[5]))
                    nomes_clientes.append(row[0])  # Adiciona o nome do cliente

            # Calcular comissão de 6%
            comissoes = [valor * 0.06 for valor in valores_vendidos]

            # Exibir gráfico com nomes dos clientes e comissões
            plt.bar(nomes_clientes, comissoes)
            plt.title('Comissões de 6% sobre imóveis vendidos')
            plt.ylabel('Valor da Comissão (R$)')
            plt.xlabel('Clientes')
            plt.xticks(rotation=45, ha='right')
            plt.tight_layout()  # Ajustar o layout para as labels não ficarem sobrepostas
            plt.show()

        # Variáveis de texto
        name_value = StringVar()
        contact_value = StringVar()
        code_value = StringVar()
        address_value = StringVar()
        status_var = StringVar()

        # Entradas
        name_entry = ctk.CTkEntry(self, width=350, textvariable=name_value, font=("Century Gothic bold", 16), fg_color="transparent")
        contact_entry = ctk.CTkEntry(self, width=200, textvariable=contact_value, font=("Century Gothic bold", 16), fg_color="transparent")
        code_entry = ctk.CTkEntry(self, width=150, textvariable=code_value, font=("Century Gothic bold", 16), fg_color="transparent")
        address_entry = ctk.CTkEntry(self, width=200, textvariable=address_value, font=("Century Gothic bold", 16), fg_color="transparent")
        obs_entry = ctk.CTkTextbox(self, width=500, height=150, font=("arial", 18), border_color="#aaa", border_width=2, fg_color="transparent")
        value_entry = ctk.CTkEntry(self, width=200, font=("Century Gothic bold", 16), fg_color="transparent")

        # Status checkbox
        status_options = ["Vendido", "Encerrado", "Em Processo"]
        status_menu = ctk.CTkOptionMenu(self, values=status_options, variable=status_var)

        # Labels
        lb_name = ctk.CTkLabel(frame, text="Nome", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_contact = ctk.CTkLabel(frame, text="Contato", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_code = ctk.CTkLabel(frame, text="Código do Imóvel", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_address = ctk.CTkLabel(frame, text="Endereço", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_obs = ctk.CTkLabel(frame, text="Observação", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_value = ctk.CTkLabel(frame, text="Valor do Imóvel", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_status = ctk.CTkLabel(frame, text="Status", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])

        # Botões
        btn_submit = ctk.CTkButton(self, text="Salvar Dados".upper(), command=submit, fg_color="#151", hover_color="#131")
        btn_clear = ctk.CTkButton(self, text="Limpar Dados".upper(), command=clear, fg_color="#555", hover_color="#333")
        btn_calcular = ctk.CTkButton(self, text="Calcular Comissões".upper(), command=calcular_comissoes, fg_color="#007BFF", hover_color="#0056b3")

        # Posicionando as labels e entradas
        lb_name.place(x=50, y=120)
        name_entry.place(x=50, y=150)

        lb_contact.place(x=550, y=120)
        contact_entry.place(x=550, y=150)

        lb_code.place(x=50, y=190)
        code_entry.place(x=50, y=220)

        lb_address.place(x=550, y=190)
        address_entry.place(x=550, y=220)

        lb_obs.place(x=50, y=260)
        obs_entry.place(x=50, y=290)

        lb_value.place(x=550, y=260)
        value_entry.place(x=550, y=290)

        lb_status.place(x=50, y=450)
        status_menu.place(x=50, y=480)

        btn_submit.place(x=300, y=520)
        btn_clear.place(x=450, y=520)
        btn_calcular.place(x=600, y=520)

    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)


if __name__ == "__main__":
    app = App()
    app.mainloop()